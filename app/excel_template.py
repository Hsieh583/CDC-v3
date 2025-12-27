from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime


def create_procurement_template(case_number: str, title: str = "") -> BytesIO:
    """
    Create a procurement request Excel template
    建立請購單 Excel 範本
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "請購單"
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "請購單 Procurement Request"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Case info section
    ws['A3'] = "案件編號 Case Number:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = case_number
    
    ws['A4'] = "案件名稱 Title:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = title
    ws.merge_cells('B4:E4')
    
    ws['A5'] = "建立日期 Created Date:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = datetime.now().strftime("%Y-%m-%d")
    
    # Empty row
    ws.row_dimensions[6].height = 5
    
    # Item list header
    headers = ['項次\nItem', '品名/規格\nDescription', '數量\nQuantity', '單位\nUnit', '備註\nRemarks']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Add 10 empty rows for items
    for row in range(8, 18):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col == 1:  # Item number
                cell.value = row - 7
                cell.alignment = Alignment(horizontal='center')
    
    # Footer section
    ws['A19'] = "請購人 Requestor:"
    ws['A19'].font = Font(bold=True)
    ws['B19'] = ""
    
    ws['A20'] = "部門 Department:"
    ws['A20'].font = Font(bold=True)
    ws['B20'] = ""
    
    ws['A21'] = "聯絡電話 Contact:"
    ws['A21'].font = Font(bold=True)
    ws['B21'] = ""
    
    # Notes section
    ws['A23'] = "備註說明 Notes:"
    ws['A23'].font = Font(bold=True)
    ws.merge_cells('A23:E23')
    ws.merge_cells('A24:E27')
    ws['A24'].alignment = Alignment(vertical='top', wrap_text=True)
    
    # Apply border to notes section
    for row in range(24, 28):
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_blank_template() -> BytesIO:
    """
    Create a blank procurement template for download
    建立空白請購單範本供下載
    """
    return create_procurement_template("CDC-PR-XXXX-XXXXX", "請填寫案件名稱")
